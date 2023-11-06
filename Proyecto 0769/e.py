import tkinter as tk
from openpyxl import Workbook
from tkinter import messagebox

def mostrar_info():
    nombres_profesores = []
    try:
        with open("cursos.txt", "r") as file:
            lines = file.read().split('\n\n')  # Dividir por cursos
            for course_info in lines:
                nombre = None
                profesor = None
                course_lines = course_info.split('\n')
                for line in course_lines:
                    if line.startswith("Nombre: "):
                        nombre = line.split(': ')[1]
                    elif line.startswith("Profesor: "):
                        profesor = line.split(': ')[1]
                if nombre and profesor:
                    nombres_profesores.append((profesor, nombre))
    except FileNotFoundError:
        print("El archivo de cursos no se encuentra.")

    return nombres_profesores

def generar_excel():
    info = mostrar_info()
    if info:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Profesor'
        ws['B1'] = 'Curso'
        for row, data in enumerate(info, start=2):
            ws[f'A{row}'] = data[0]
            ws[f'B{row}'] = data[1]
        wb.save("informacion.xlsx")
        messagebox.showinfo("Éxito", "Descarga exitosa. Archivo Excel generado.")
        print("Archivo Excel generado.")

def mostrar_interfaz():
    root = tk.Tk()
    root.title("Información de Cursos")
    root.geometry('500x500')
    root.configure(bg='white')

    info_frame = tk.Frame(root, bg='lightgrey', padx=10, pady=10)
    info_frame.pack(fill='both', expand=True)

    info = mostrar_info()

    for data in info:
        label = tk.Label(info_frame, text=f"{data[0]} - {data[1]}", anchor='w', justify='left')
        label.pack(fill='x')

    button_descargar = tk.Button(root, text="Descargar información", command=generar_excel)
    button_descargar.pack(side='left', padx=10, pady=10)

    button_cerrar = tk.Button(root, text="Cerrar", command=root.destroy)
    button_cerrar.pack(side='right', padx=10, pady=10)

    root.mainloop()

mostrar_interfaz()